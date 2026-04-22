"""
Production-ready RAG SOW extractor: Vertex AI (Gemini) + SharePoint + Smart Resume.
Uses same connection pattern as check_connections.py (SHAREPOINT_DRIVE_ID, _resolve_drive_id, etc.).
Entry point: run with uvicorn or via Docker.
"""
import asyncio
import json
import logging
import os
import threading
from contextlib import asynccontextmanager
from io import BytesIO
from pathlib import Path
from queue import Empty, Queue
from typing import Callable, Dict, Optional

from dotenv import load_dotenv

# Load baked-in or local .env next to this package (e.g. /app/.env in Docker).
load_dotenv(Path(__file__).resolve().parent / ".env")
# Optional: load SharePoint secrets from GCP Secret Manager into os.environ (before config is read)
try:
    from secret_loader import load_secrets_from_gcp
    load_secrets_from_gcp()
except Exception:  # noqa: S110
    pass

import pandas as pd
from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from starlette.middleware.sessions import SessionMiddleware

from config import (
    DATA_DIR,
    EXCEL_OUTPUT_DIR,
    EXCEL_SOW_PATH,
    EXCEL_INVOICE_PATH,
    EXTRACTION_MONITOR_INTERVAL_MINUTES,
    EXTRACTION_STATE_PATH,
    FIELDS,
    INVOICE_FIELDS,
    SOW_FIELDS,
    GCS_OUTPUT_BUCKET,
)
from data_utils import (
    clean_amount,
    clean_date,
    dedupe_by_contract_id,
    dedupe_keep_latest_revision,
    remove_duplicate_entries,
)
from state_manager import (
    classify_item,
    get_processed_ids,
    get_processed_meta,
    mark_processed,
    record_dedup_decisions,
    should_process_item,
)
from sharepoint_utils import (
    is_sharepoint_configured,
    browse_site_contents,
    list_contents_at_path,
    list_pdf_items_streaming,
    download_file_bytes,
    verify_sharepoint_path_reachable,
)
# Single-shot extraction: root ai_processor.py only (Gemini OCR + JSON extraction).
from ai_processor import _normalize_field_value, debug_pdf_bytes, process_pdf_bytes
from auth_utils import (
    auth_start_url,
    create_login_state,
    exchange_code_for_token,
    get_current_access_token,
    get_current_user,
)
from user_storage import user_blob_prefix, user_paths

# Internal key used to carry the SharePoint web URL through the result row dict.
# Written as a data column "SharePoint URL" in both Excels so Filename hyperlinks
# can be re-applied when the Excel is reloaded on subsequent runs.
_URL_KEY = "SharePoint URL"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@asynccontextmanager
async def _app_lifespan(app: FastAPI):
    """Startup hook. Per-user state sync happens inside user-scoped extraction calls."""
    yield


app = FastAPI(
    title="SOW RAG Extractor (Vertex AI + SharePoint)",
    lifespan=_app_lifespan,
)
app.add_middleware(
    SessionMiddleware,
    secret_key=os.environ.get("SESSION_SECRET", "replace-me-session-secret"),
    same_site="lax",
    https_only=False,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in os.environ.get("CORS_ALLOWED_ORIGINS", "*").split(",") if o.strip()],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATA_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Background extraction state (same connection approach as check_connections; Smart Resume = state file)
_extraction_lock = threading.Lock()
_extraction_state = {
    "running": False,
    "stop_requested": False,
    "current_file": None,
    "processed_this_run": 0,
    "total_to_process": 0,
    "last_error": None,
    "total_in_excel": 0,
}
_user_extraction_state: Dict[str, dict] = {}


def _user_id_from_claims(user: dict) -> str:
    return (user.get("oid") or user.get("sub") or user.get("preferred_username") or "anonymous").strip()


def _default_user_state() -> dict:
    return {
        "running": False,
        "stop_requested": False,
        "current_file": None,
        "processed_this_run": 0,
        "total_to_process": 0,
        "last_error": None,
        "total_in_excel": 0,
    }


def _get_user_state(user_id: str) -> dict:
    with _extraction_lock:
        if user_id not in _user_extraction_state:
            _user_extraction_state[user_id] = _default_user_state()
        return _user_extraction_state[user_id]


def _load_user_context(user_id: str) -> dict:
    p = user_paths(user_id)["context_path"]
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_user_context(user_id: str, context: dict) -> None:
    p = user_paths(user_id)["context_path"]
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(context, indent=2), encoding="utf-8")


def _monitor_loop() -> None:
    """Disabled for delegated per-user mode: monitor cannot run without a user context/token."""
    import time
    interval_sec = EXTRACTION_MONITOR_INTERVAL_MINUTES * 60
    if interval_sec <= 0:
        return
    logger.warning(
        "Background monitor is disabled in delegated-user mode. "
        "Use authenticated UI/API calls per user to run extraction."
    )
    while True:
        time.sleep(interval_sec)
        continue


def _start_monitor_if_enabled() -> None:
    if EXTRACTION_MONITOR_INTERVAL_MINUTES <= 0:
        return
    t = threading.Thread(target=_monitor_loop, daemon=True)
    t.start()


# Start/End dates are stored as extracted (original document format); not normalized to ISO.
_AMOUNT_COLS = {"Commercial Value", "TCV", "Annual Value"}


def _load_existing_excel(excel_path: Path, output_fields: list) -> Optional[pd.DataFrame]:
    """
    Load an existing Excel into a DataFrame with columns [Filename, SharePoint URL, *output_fields].
    Returns None if the file doesn't exist or can't be read.
    Also migrates legacy 'SharePoint Link' column to 'SharePoint URL' transparently.
    """
    if not excel_path.exists():
        return None
    try:
        df = pd.read_excel(excel_path)
        # Migrate old column name
        if _URL_KEY not in df.columns and "SharePoint Link" in df.columns:
            df = df.rename(columns={"SharePoint Link": _URL_KEY})
        for col in ["Filename", _URL_KEY] + output_fields:
            if col not in df.columns:
                df[col] = ""
        keep = ["Filename", _URL_KEY] + output_fields + (["Error"] if "Error" in df.columns else [])
        return df[[c for c in keep if c in df.columns]].copy()
    except Exception as e:
        logger.warning("Could not load %s: %s", excel_path.name, e)
        return None


def _apply_filename_hyperlinks(path: Path) -> None:
    """
    Post-process the Excel: read 'SharePoint URL' column, set each Filename cell as a
    clickable hyperlink, then hide the URL column (set width=0 / very narrow).
    Filename is styled in blue underline to signal it is clickable.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    try:
        wb = load_workbook(str(path))
        ws = wb.active
        if not ws or ws.max_row < 2:
            return
        header = [cell.value for cell in ws[1]]
        try:
            fn_idx = header.index("Filename") + 1
            url_idx = header.index(_URL_KEY) + 1
        except ValueError:
            return
        hyperlink_font = Font(color="0563C1", underline="single")
        for row_num in range(2, ws.max_row + 1):
            fn_cell = ws.cell(row=row_num, column=fn_idx)
            url_cell = ws.cell(row=row_num, column=url_idx)
            url = (str(url_cell.value or "")).strip()
            if url and (url.startswith("http://") or url.startswith("https://")):
                fn_cell.hyperlink = url
                fn_cell.font = hyperlink_font
        # Make the URL column very narrow (keep data for reload but visually minimal)
        ws.column_dimensions[get_column_letter(url_idx)].width = 4
        wb.save(str(path))
    except Exception as e:
        logger.warning("Could not apply filename hyperlinks to %s: %s", path.name, e)


def _write_excel(
    excel_path: Path,
    existing_df: Optional[pd.DataFrame],
    new_results: list,
    output_fields: list,
    gcs_blob_name: Optional[str] = None,
    upload_to_gcs: bool = True,
    state_path: Optional[Path] = None,
) -> int:
    """
    Build cumulative DataFrame (existing + new rows), dedupe, write to *excel_path*.
    Columns: Filename | SharePoint URL | *output_fields* | Error
    Filename cells are hyperlinked to SharePoint URL after writing.
    Optionally uploads the Excel to GCS.  Returns total row count.
    """
    logger.info("Saving %s (%s new rows)...", excel_path.name, len(new_results))
    base_cols = ["Filename", _URL_KEY] + output_fields

    if existing_df is not None and not existing_df.empty:
        df = existing_df.copy()
        for c in base_cols:
            if c not in df.columns:
                df[c] = ""
        df = df[[c for c in base_cols + ["Error"] if c in df.columns]].copy()
    else:
        df = pd.DataFrame(columns=base_cols)

    if new_results:
        new_df = pd.DataFrame(new_results)
        for col in base_cols + ["Error"]:
            if col not in new_df.columns:
                new_df[col] = ""
        new_df = new_df[[c for c in base_cols + ["Error"] if c in new_df.columns]].copy()
        for col in new_df.columns:
            if col in _AMOUNT_COLS:
                new_df[col] = new_df[col].apply(clean_amount)
        # Replace stale rows when the same file is reprocessed.
        if not df.empty:
            new_keys = set(
                zip(
                    new_df["Filename"].astype(str).str.strip(),
                    new_df[_URL_KEY].astype(str).str.strip(),
                )
            )
            if new_keys:
                existing_keys = list(
                    zip(
                        df["Filename"].astype(str).str.strip(),
                        df[_URL_KEY].astype(str).str.strip(),
                    )
                )
                keep_mask = [k not in new_keys for k in existing_keys]
                replaced = len(df) - sum(keep_mask)
                if replaced > 0:
                    logger.info(
                        "Replacing %s existing row(s) for reprocessed files in %s.",
                        replaced,
                        excel_path.name,
                    )
                df = df[keep_mask].copy()
        df = pd.concat([df, new_df], ignore_index=True)

    for date_col, position in (("Start Date", "start"), ("End Date", "end")):
        if date_col in df.columns:
            df[date_col] = df[date_col].apply(lambda v: clean_date(v, month_year_position=position))

    # Same filename base with revision markers → keep only the latest revision
    df, rev_dedup_log = dedupe_keep_latest_revision(df, output_fields)
    # Same base Contract ID (e.g. 4533232908 vs 4533232908-1) → keep latest version by Start Date
    df, cid_dedup_log = dedupe_by_contract_id(df)
    # Persist dedup decisions to extraction_state.json
    all_dedup_log = rev_dedup_log + cid_dedup_log
    if all_dedup_log:
        record_dedup_decisions(all_dedup_log, state_path=state_path)
    # Exact duplicate rows (e.g. same file processed twice); include Filename so rev variants are not collapsed here
    df = remove_duplicate_entries(df, fields=output_fields + ["Filename"])
    df.to_excel(str(excel_path), index=False)
    _apply_filename_hyperlinks(excel_path)
    logger.info("%s saved (%s rows).", excel_path.name, len(df))

    if upload_to_gcs and GCS_OUTPUT_BUCKET:
        try:
            from gcs_utils import upload_file_to_bucket
            blob = gcs_blob_name or excel_path.name
            gs_url = upload_file_to_bucket(excel_path, GCS_OUTPUT_BUCKET, blob_name=blob)
            logger.info("%s uploaded to GCS: %s", excel_path.name, gs_url)
        except Exception as e:
            logger.exception("GCS upload failed for %s: %s", excel_path.name, e)

    return len(df)


_STATE_BLOB_NAME = "extraction_state.json"


def _sync_persistent_data_from_gcs(
    state_path: Path = EXTRACTION_STATE_PATH,
    excel_sow_path: Path = EXCEL_SOW_PATH,
    excel_invoice_path: Path = EXCEL_INVOICE_PATH,
    blob_prefix: str = "",
) -> None:
    """
    Download extraction_state.json and cumulative Excel files from GCS when configured.

    Needed for Cloud Run (ephemeral disk): a new instance must restore state and Excels
    before Smart Resume and /download work. Safe no-op when GCS_OUTPUT_BUCKET is unset.
    """
    if not GCS_OUTPUT_BUCKET:
        return
    try:
        from gcs_utils import download_file_from_bucket
        state_blob = "/".join([p for p in [blob_prefix.strip("/"), _STATE_BLOB_NAME] if p])
        ok = download_file_from_bucket(GCS_OUTPUT_BUCKET, state_blob, state_path)
        if ok:
            logger.info("Extraction state synced from GCS (gs://%s/%s).", GCS_OUTPUT_BUCKET, state_blob)
        else:
            logger.info("No extraction state in GCS; using local file if present.")
        for excel_path in (excel_sow_path, excel_invoice_path):
            blob_name = "/".join([p for p in [blob_prefix.strip("/"), excel_path.name] if p])
            if download_file_from_bucket(GCS_OUTPUT_BUCKET, blob_name, excel_path):
                logger.info("Excel synced from GCS: gs://%s/%s", GCS_OUTPUT_BUCKET, blob_name)
            else:
                logger.info("No %s in GCS yet; will create on first successful save.", blob_name)
    except Exception as e:
        logger.warning("Could not sync persistent data from GCS: %s. Using local files if present.", e)


def _upload_state_to_gcs(
    state_path: Path = EXTRACTION_STATE_PATH,
    blob_prefix: str = "",
) -> None:
    """Upload extraction_state.json to GCS alongside the Excel files."""
    if not GCS_OUTPUT_BUCKET or not state_path.exists():
        return
    try:
        from gcs_utils import upload_file_to_bucket
        state_blob = "/".join([p for p in [blob_prefix.strip("/"), _STATE_BLOB_NAME] if p])
        gs_url = upload_file_to_bucket(state_path, GCS_OUTPUT_BUCKET, blob_name=state_blob)
        logger.info("Extraction state uploaded to GCS: %s", gs_url)
    except Exception as e:
        logger.warning("Failed to upload extraction state to GCS: %s", e)


def _route_doc_type(doc_type_raw: str) -> str:
    """Return 'invoice' or 'sow' from a raw Document Type string extracted by the LLM."""
    v = doc_type_raw.strip().lower()
    if "invoice" in v or "licen" in v:
        return "invoice"
    return "sow"


def _run_extraction(
    stop_check: Optional[Callable[[], bool]] = None,
    force_reprocess: bool = False,
    user_id: str = "global",
    access_token: Optional[str] = None,
    user_context: Optional[dict] = None,
):
    """
    Stream PDFs from SharePoint (listing in background); process each as soon as it arrives.
    Routes each result to sow_results.xlsx or invoice_results.xlsx based on detected Document Type.
    Unless force_reprocess=True: skip already-processed (Smart Resume) and only process new/changed.
    Saves both Excels periodically and on stop/finish.
    Returns (new_count, total_sow + total_invoice, [EXCEL_SOW_PATH, EXCEL_INVOICE_PATH]).
    """
    if not is_sharepoint_configured():
        logger.warning("SharePoint not configured; no PDF source")
        return 0, 0, None

    ustate = _get_user_state(user_id)
    with _extraction_lock:
        ustate["current_file"] = "Listing + processing (starting as PDFs are found)..."
        ustate["total_to_process"] = -1
    ctx = user_context or _load_user_context(user_id)
    site_url = (ctx.get("site_url") or "").strip()
    drive_id_override = (ctx.get("drive_id") or "").strip() or None
    drive_path = (ctx.get("drive_path") or "").strip()
    if not site_url:
        raise RuntimeError("User context is missing SharePoint site_url. Set it via /sharepoint/context first.")
    upaths = user_paths(user_id)
    state_path = upaths["state_path"]
    excel_sow_path = upaths["excel_sow_path"]
    excel_invoice_path = upaths["excel_invoice_path"]
    blob_prefix = user_blob_prefix(user_id)

    _sync_persistent_data_from_gcs(
        state_path=state_path,
        excel_sow_path=excel_sow_path,
        excel_invoice_path=excel_invoice_path,
        blob_prefix=blob_prefix,
    )

    logger.info(
        "Extraction started: listing and processing in parallel (%s). "
        "Routes each PDF to sow_results.xlsx or invoice_results.xlsx by document type.",
        "reprocess all (ignoring state)" if force_reprocess else "new + changed files",
    )
    processed_ids = get_processed_ids(state_path=state_path)
    processed_meta = get_processed_meta(state_path=state_path)
    existing_sow_df = _load_existing_excel(excel_sow_path, SOW_FIELDS)
    existing_inv_df = _load_existing_excel(excel_invoice_path, INVOICE_FIELDS)

    pdf_queue: "Queue[Optional[tuple]]" = Queue()

    def producer() -> None:
        try:
            for item, drive_id in list_pdf_items_streaming(
                access_token=access_token,
                site_url=site_url,
                drive_id=drive_id_override,
                drive_path=drive_path,
            ):
                if stop_check and stop_check():
                    logger.info("Stop requested; stopping PDF listing.")
                    break
                pdf_queue.put((item, drive_id))
        except Exception as e:
            logger.exception("Listing failed: %s", e)
        finally:
            pdf_queue.put(None)

    list_thread = threading.Thread(target=producer, daemon=True)
    list_thread.start()

    sow_results: list = []
    invoice_results: list = []
    processed_this_batch = 0
    SAVE_EXCEL_EVERY_N = 10  # periodic local saves during long runs (GCS upload at end/stop only)

    def _periodic_save() -> None:
        _write_excel(excel_sow_path, existing_sow_df, sow_results, SOW_FIELDS, upload_to_gcs=False, state_path=state_path)
        _write_excel(excel_invoice_path, existing_inv_df, invoice_results, INVOICE_FIELDS, upload_to_gcs=False, state_path=state_path)
        logger.info(
            "Periodic save: %s SOW rows, %s Invoice rows (GCS upload at end of run).",
            len(sow_results), len(invoice_results),
        )

    def process_one(it: dict, drive_id: str, idx: int) -> None:
        nonlocal processed_this_batch
        short_name = it.get("name") or it.get("id")
        folder_name = it.get("folder", "")
        with _extraction_lock:
            ustate["current_file"] = short_name
            ustate["last_error"] = None
        logger.info("[%s] Downloading: %s (%s)", idx, short_name, folder_name or "root")
        try:
            pdf_bytes = download_file_bytes(it["id"], drive_id=drive_id, access_token=access_token, site_url=site_url)
            size_kb = len(pdf_bytes) / 1024
            logger.info("[%s] Downloaded %s (%.1f KB). Extracting fields (single-shot)...", idx, short_name, size_kb)
            rows = process_pdf_bytes(pdf_bytes, folder_name=folder_name, file_name=short_name)
            # process_pdf_bytes now returns a list of row dicts (multi-row for multi-table invoices)
            first_row = rows[0] if rows else {}
            doc_type = (first_row.get("Document Type") or "").strip()
            target = _route_doc_type(doc_type)
            for row in rows:
                debug_note = row.pop("_debug_note", None)
                if debug_note:
                    row["Error"] = debug_note
                row[_URL_KEY] = it.get("webUrl", "")
                if target == "invoice":
                    invoice_results.append(row)
                else:
                    sow_results.append(row)
            row_count = len(rows)
            logger.info(
                "[%s] Completed: %s → %s (%s row%s, Type: %s)",
                idx, short_name, "invoice_results" if target == "invoice" else "sow_results",
                row_count, "" if row_count == 1 else "s", doc_type or "unknown",
            )
            mark_processed(
                it["id"],
                eTag=it.get("eTag"),
                last_modified=it.get("last_modified"),
                name=it.get("name"),
                folder=it.get("folder"),
                doc_type=target,
                original_language=(first_row.get("Original Language") or "").strip() or None,
                state_path=state_path,
            )
            processed_this_batch += 1
            with _extraction_lock:
                ustate["processed_this_run"] = processed_this_batch
            if processed_this_batch % SAVE_EXCEL_EVERY_N == 0:
                _periodic_save()
        except Exception as e:
            logger.exception("[%s] FAILED: %s — %s", idx, short_name, e)
            with _extraction_lock:
                ustate["last_error"] = str(e)
            err_row = {
                "Filename": it.get("name", ""),
                _URL_KEY: it.get("webUrl", ""),
                "Error": str(e),
                **{f: "" for f in FIELDS},
            }
            sow_results.append(err_row)

    def _save_and_return(upload: bool = True):
        sow_total = _write_excel(
            excel_sow_path, existing_sow_df, sow_results, SOW_FIELDS,
            gcs_blob_name=f"{blob_prefix}/{excel_sow_path.name}", upload_to_gcs=upload,
            state_path=state_path,
        )
        inv_total = _write_excel(
            excel_invoice_path, existing_inv_df, invoice_results, INVOICE_FIELDS,
            gcs_blob_name=f"{blob_prefix}/{excel_invoice_path.name}", upload_to_gcs=upload,
            state_path=state_path,
        )
        if upload:
            _upload_state_to_gcs(state_path=state_path, blob_prefix=blob_prefix)
        return sow_total + inv_total

    idx = 0
    while True:
        if stop_check and stop_check():
            logger.info("Stop requested; saving both Excels and uploading to GCS before exiting.")
            total_count = _save_and_return(upload=True)
            return len(sow_results) + len(invoice_results), total_count, excel_sow_path
        x = pdf_queue.get()
        if x is None:
            break
        item, drive_id = x
        idx += 1
        if not force_reprocess and not should_process_item(item, processed_ids, processed_meta):
            continue
        process_one(item, drive_id, idx)

    while True:
        try:
            x = pdf_queue.get_nowait()
        except Empty:
            break
        if x is None:
            continue
        item, drive_id = x
        idx += 1
        if not force_reprocess and not should_process_item(item, processed_ids, processed_meta):
            continue
        process_one(item, drive_id, idx)

    if idx == 0:
        logger.warning("No PDFs found at path. Check SHAREPOINT_DRIVE_PATH and SHAREPOINT_DRIVE_ID.")
        with _extraction_lock:
            ustate["current_file"] = "No PDFs found at path (check selected site/path/drive)"
            ustate["total_to_process"] = 0
        total = (
            (len(existing_sow_df) if existing_sow_df is not None else 0)
            + (len(existing_inv_df) if existing_inv_df is not None else 0)
        )
        return 0, total, None

    new_total = len(sow_results) + len(invoice_results)
    logger.info(
        "Extraction run finished. %s new: %s SOW, %s Invoice. Saving and uploading...",
        new_total, len(sow_results), len(invoice_results),
    )
    total_count = _save_and_return(upload=True)
    logger.info("Done. Combined rows across both Excels: %s.", total_count)
    return new_total, total_count, excel_sow_path


def _run_extraction_background(
    user_id: str,
    access_token: str,
    user_context: dict,
    force_reprocess: bool = False,
) -> None:
    """Run extraction in background; updates _extraction_state. Uses stop_check so stop saves Excel + GCS."""
    ustate = _get_user_state(user_id)

    def stop_check() -> bool:
        with _extraction_lock:
            return ustate["stop_requested"]

    logger.info("Background extraction job started (force_reprocess=%s).", force_reprocess)
    try:
        with _extraction_lock:
            ustate["processed_this_run"] = 0
            ustate["last_error"] = None
        new_count, total_count, output_path = _run_extraction(
            stop_check=stop_check,
            force_reprocess=force_reprocess,
            user_id=user_id,
            access_token=access_token,
            user_context=user_context,
        )
        with _extraction_lock:
            ustate["total_in_excel"] = total_count
        logger.info("Background extraction finished successfully. New: %s, total in Excel: %s", new_count, total_count)
    except Exception as e:
        logger.exception("Background extraction failed: %s", e)
        with _extraction_lock:
            ustate["last_error"] = str(e)
    finally:
        with _extraction_lock:
            ustate["running"] = False
            ustate["stop_requested"] = False
            ustate["current_file"] = None
        logger.info("Extraction job ended. Check /extract-sow/status or /state for summary.")


@app.get("/auth/login")
async def auth_login(request: Request):
    """Start delegated Microsoft OAuth2 login."""
    state = create_login_state()
    request.session["oauth_state"] = state
    return RedirectResponse(auth_start_url(request, state=state))


@app.get("/auth/callback")
async def auth_callback(request: Request, code: Optional[str] = None, state: Optional[str] = None):
    expected = request.session.get("oauth_state")
    if not code or not state or not expected or state != expected:
        raise HTTPException(status_code=401, detail="Invalid OAuth callback state.")
    token_result = exchange_code_for_token(request, code=code)
    claims = token_result.get("id_token_claims") or {}
    request.session["user"] = {
        "oid": claims.get("oid") or claims.get("sub"),
        "name": claims.get("name") or claims.get("preferred_username") or "User",
        "preferred_username": claims.get("preferred_username") or "",
    }
    request.session["tokens"] = {
        "access_token": token_result["access_token"],
        "refresh_token": token_result.get("refresh_token"),
        "expires_at": int(__import__("time").time()) + int(token_result.get("expires_in", 3600)),
        "id_token_claims": claims,
    }
    request.session.pop("oauth_state", None)
    return RedirectResponse(url="/")


@app.post("/auth/logout")
async def auth_logout(request: Request):
    request.session.clear()
    return {"status": "ok", "message": "Logged out"}


@app.get("/auth/me")
async def auth_me(request: Request):
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user


@app.get("/sharepoint/context")
async def get_sharepoint_context(user: dict = Depends(get_current_user)):
    user_id = _user_id_from_claims(user)
    return _load_user_context(user_id)


@app.post("/sharepoint/context")
async def set_sharepoint_context(
    request: Request,
    payload: dict,
    user: dict = Depends(get_current_user),
):
    user_id = _user_id_from_claims(user)
    site_url = (payload.get("site_url") or "").strip()
    drive_path = (payload.get("drive_path") or "").strip().strip("/")
    drive_id = (payload.get("drive_id") or "").strip()
    if not site_url:
        raise HTTPException(status_code=400, detail="site_url is required.")
    access_token = get_current_access_token(request)
    ok, msg = verify_sharepoint_path_reachable(
        token=access_token,
        site_url=site_url,
        drive_id=drive_id or None,
        folder_path=drive_path,
    )
    if not ok:
        raise HTTPException(status_code=400, detail=f"SharePoint context validation failed: {msg}")
    ctx = {"site_url": site_url, "drive_path": drive_path, "drive_id": drive_id}
    _save_user_context(user_id, ctx)
    return {"status": "ok", "context": ctx, "validation": msg}


@app.get("/sharepoint/browse")
async def browse_sharepoint(
    request: Request,
    folder_path: str = "",
    user: dict = Depends(get_current_user),
):
    user_id = _user_id_from_claims(user)
    ctx = _load_user_context(user_id)
    site_url = (ctx.get("site_url") or "").strip()
    if not site_url:
        raise HTTPException(status_code=400, detail="SharePoint context not set. Call /sharepoint/context first.")
    access_token = get_current_access_token(request)
    items = list_contents_at_path(
        folder_path=folder_path,
        token=access_token,
        drive_id=(ctx.get("drive_id") or "").strip() or None,
        site_url=site_url,
    )
    return {"items": items, "count": len(items)}


@app.post("/extract-sow/")
async def extract_sow(
    request: Request,
    force_reprocess: bool = False,
    user: dict = Depends(get_current_user),
):
    """
    Run extraction synchronously (blocks until done). Uses Smart Resume unless force_reprocess=true.
    Query: force_reprocess=true to process every file from start (ignore state).
    For background run: POST /extract-sow/start or POST /extract-sow/reprocess-all.
    """
    try:
        user_id = _user_id_from_claims(user)
        access_token = get_current_access_token(request)
        ctx = _load_user_context(user_id)
        if not ctx.get("site_url"):
            raise HTTPException(status_code=400, detail="SharePoint context not set. Call /sharepoint/context first.")
        loop = asyncio.get_event_loop()
        new_count, total_count, output_path = await loop.run_in_executor(
            None,
            lambda: _run_extraction(
                stop_check=None,
                force_reprocess=force_reprocess,
                user_id=user_id,
                access_token=access_token,
                user_context=ctx,
            ),
        )
    except Exception as e:
        logger.exception("Extract failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

    if output_path is None:
        return {
            "status": "success",
            "message": "No PDFs found; no Excel written.",
            "output_file": None,
            "new_records": 0,
            "total_records": 0,
        }

    return {
        "status": "success",
        "output_file": output_path.name,
        "new_records": new_count,
        "total_records": total_count,
    }


@app.post("/extract-sow/start")
async def extract_sow_start(request: Request, user: dict = Depends(get_current_user)):
    """Start extraction in background. Resumes from state (skips already-processed PDFs). Use GET /extract-sow/status to poll; POST /extract-sow/stop to stop."""
    user_id = _user_id_from_claims(user)
    ctx = _load_user_context(user_id)
    if not ctx.get("site_url"):
        raise HTTPException(status_code=400, detail="SharePoint context not set. Call /sharepoint/context first.")
    access_token = get_current_access_token(request)
    ustate = _get_user_state(user_id)
    with _extraction_lock:
        if ustate["running"]:
            return {"status": "already_running", "message": "Extraction is already running."}
        ustate["running"] = True
        ustate["stop_requested"] = False
        ustate["processed_this_run"] = 0
        ustate["total_to_process"] = 0
        ustate["current_file"] = None
        ustate["last_error"] = None
    thread = threading.Thread(
        target=_run_extraction_background,
        kwargs={"user_id": user_id, "access_token": access_token, "user_context": ctx, "force_reprocess": False},
        daemon=True,
    )
    thread.start()
    return {"status": "started", "message": "Extraction started in background. Use GET /extract-sow/status to poll; POST /extract-sow/stop to stop."}


@app.post("/extract-sow/reprocess-all")
async def extract_sow_reprocess_all(request: Request, user: dict = Depends(get_current_user)):
    """
    Start extraction in background and process every PDF from start, ignoring state (already-processed files are reprocessed).
    Use GET /extract-sow/status to poll; POST /extract-sow/stop to stop.
    """
    user_id = _user_id_from_claims(user)
    ctx = _load_user_context(user_id)
    if not ctx.get("site_url"):
        raise HTTPException(status_code=400, detail="SharePoint context not set. Call /sharepoint/context first.")
    access_token = get_current_access_token(request)
    ustate = _get_user_state(user_id)
    with _extraction_lock:
        if ustate["running"]:
            return {"status": "already_running", "message": "Extraction is already running."}
        ustate["running"] = True
        ustate["stop_requested"] = False
        ustate["processed_this_run"] = 0
        ustate["total_to_process"] = 0
        ustate["current_file"] = None
        ustate["last_error"] = None
    thread = threading.Thread(
        target=_run_extraction_background,
        kwargs={"user_id": user_id, "access_token": access_token, "user_context": ctx, "force_reprocess": True},
        daemon=True,
    )
    thread.start()
    return {
        "status": "started",
        "message": "Reprocess-all started: every PDF will be processed from start (state ignored). Use GET /extract-sow/status to poll; POST /extract-sow/stop to stop.",
    }


@app.get("/extract-sow/status")
async def extract_sow_status(user: dict = Depends(get_current_user)):
    """Return current extraction status: running, stop_requested, progress (processed, total, current_file), last_error."""
    user_id = _user_id_from_claims(user)
    with _extraction_lock:
        s = dict(_get_user_state(user_id))
    # Helpful message when running: show phase (listing vs processing) or why 0 processed
    phase = "idle"
    if s["running"]:
        if s["current_file"] and "Listing" in str(s["current_file"]):
            phase = "listing"
        elif s["total_to_process"] and s["total_to_process"] > 0:
            phase = "processing"
        else:
            phase = "starting"
    return {
        "running": s["running"],
        "stop_requested": s["stop_requested"],
        "phase": phase,
        "processed_this_run": s["processed_this_run"],
        "total_to_process": s["total_to_process"],
        "current_file": s["current_file"],
        "total_in_excel": s["total_in_excel"],
        "last_error": s["last_error"],
    }


@app.post("/extract-sow/stop")
async def extract_sow_stop(user: dict = Depends(get_current_user)):
    """Request extraction to stop after the current PDF. Excel is saved and uploaded to GCS before exit; next run resumes from state."""
    user_id = _user_id_from_claims(user)
    ustate = _get_user_state(user_id)
    with _extraction_lock:
        if not ustate["running"]:
            return {"status": "not_running", "message": "No extraction is running."}
        ustate["stop_requested"] = True
    return {"status": "stop_requested", "message": "Extraction will stop after current PDF. Excel will be saved and uploaded to GCS."}


def _run_scan(timeout_seconds: int = 300) -> dict:
    """List PDFs from SharePoint (streaming) and return new/changed/up_to_date counts. Does not process."""
    if not is_sharepoint_configured():
        return {"error": "SharePoint not configured", "new_count": 0, "changed_count": 0, "up_to_date_count": 0, "total": 0}
    import time
    processed_ids = get_processed_ids()
    processed_meta = get_processed_meta()
    scan_queue: "Queue[Optional[tuple]]" = Queue()
    def producer() -> None:
        try:
            for item, _ in list_pdf_items_streaming():
                scan_queue.put((item, None))
        except Exception as e:
            logger.exception("Scan listing failed: %s", e)
        finally:
            scan_queue.put(None)
    t = threading.Thread(target=producer, daemon=True)
    t.start()
    deadline = time.monotonic() + timeout_seconds
    new_count = changed_count = up_to_date_count = 0
    while time.monotonic() < deadline:
        try:
            x = scan_queue.get(timeout=1)
        except Empty:
            continue
        if x is None:
            break
        item, _ = x
        kind = classify_item(item, processed_ids, processed_meta)
        if kind == "new":
            new_count += 1
        elif kind == "changed":
            changed_count += 1
        else:
            up_to_date_count += 1
    # Drain remainder if we hit timeout
    while True:
        try:
            x = scan_queue.get_nowait()
        except Empty:
            break
        if x is None:
            continue
        item, _ = x
        kind = classify_item(item, processed_ids, processed_meta)
        if kind == "new":
            new_count += 1
        elif kind == "changed":
            changed_count += 1
        else:
            up_to_date_count += 1
    total = new_count + changed_count + up_to_date_count
    return {
        "new_count": new_count,
        "changed_count": changed_count,
        "up_to_date_count": up_to_date_count,
        "total": total,
        "to_process": new_count + changed_count,
        "scan_timed_out": time.monotonic() >= deadline and t.is_alive(),
    }


@app.get("/extract-sow/scan")
async def extract_sow_scan(request: Request, timeout: int = 300, user: dict = Depends(get_current_user)):
    """
    Scan SharePoint for PDFs and return counts: new, changed (modified since last run), up_to_date.
    Does not process; use POST /extract-sow/start to process new and changed files.
    Optional query: timeout= seconds to wait for listing (default 300).
    """
    try:
        user_id = _user_id_from_claims(user)
        ctx = _load_user_context(user_id)
        if not ctx.get("site_url"):
            raise HTTPException(status_code=400, detail="SharePoint context not set. Call /sharepoint/context first.")
        access_token = get_current_access_token(request)
        state_path = user_paths(user_id)["state_path"]
        processed_ids = get_processed_ids(state_path=state_path)
        processed_meta = get_processed_meta(state_path=state_path)

        def _run_user_scan(timeout_seconds: int = 300) -> dict:
            import time

            scan_queue: "Queue[Optional[tuple]]" = Queue()

            def producer() -> None:
                try:
                    for item, _ in list_pdf_items_streaming(
                        access_token=access_token,
                        site_url=ctx.get("site_url", ""),
                        drive_id=(ctx.get("drive_id") or "").strip() or None,
                        drive_path=ctx.get("drive_path", ""),
                    ):
                        scan_queue.put((item, None))
                except Exception as e:
                    logger.exception("Scan listing failed: %s", e)
                finally:
                    scan_queue.put(None)

            t = threading.Thread(target=producer, daemon=True)
            t.start()
            deadline = time.monotonic() + timeout_seconds
            new_count = changed_count = up_to_date_count = 0
            while time.monotonic() < deadline:
                try:
                    x = scan_queue.get(timeout=1)
                except Empty:
                    continue
                if x is None:
                    break
                item, _ = x
                kind = classify_item(item, processed_ids, processed_meta)
                if kind == "new":
                    new_count += 1
                elif kind == "changed":
                    changed_count += 1
                else:
                    up_to_date_count += 1
            total = new_count + changed_count + up_to_date_count
            return {
                "new_count": new_count,
                "changed_count": changed_count,
                "up_to_date_count": up_to_date_count,
                "total": total,
                "to_process": new_count + changed_count,
                "scan_timed_out": time.monotonic() >= deadline and t.is_alive(),
            }

        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, lambda: _run_user_scan(timeout_seconds=min(timeout, 600)))
    except Exception as e:
        logger.exception("Scan failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))
    return result


@app.get("/download/{file_name}")
async def download_file(file_name: str, user: dict = Depends(get_current_user)):
    """Serve a generated Excel from DATA_DIR by exact filename (e.g. sow_results.xlsx or invoice_results.xlsx)."""
    user_id = _user_id_from_claims(user)
    ups = user_paths(user_id)
    allowed = {
        ups["excel_sow_path"].name: ups["excel_sow_path"],
        ups["excel_invoice_path"].name: ups["excel_invoice_path"],
    }
    path = allowed.get(file_name)
    if path is None:
        raise HTTPException(status_code=404, detail=f"File not found: {file_name}")
    if not path.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {file_name}")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/download/filtered/{file_name}")
async def download_filtered_file(
    file_name: str,
    start_date: Optional[str] = Query(default=None, description="Include rows with Start Date on/after this date (YYYY-MM-DD)."),
    end_date: Optional[str] = Query(default=None, description="Include rows with End Date on/before this date (YYYY-MM-DD)."),
    user: dict = Depends(get_current_user),
):
    """
    Download a filtered Excel by date range.
    Filters on Start Date (>= start_date) and End Date (<= end_date) when provided.
    """
    user_id = _user_id_from_claims(user)
    ups = user_paths(user_id)
    allowed = {
        ups["excel_sow_path"].name: ups["excel_sow_path"],
        ups["excel_invoice_path"].name: ups["excel_invoice_path"],
    }
    path = allowed.get(file_name)
    if path is None:
        raise HTTPException(status_code=404, detail=f"File not found: {file_name}")
    if not path.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {file_name}")
    try:
        df = pd.read_excel(path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not read Excel: {e}")

    # Keep original if expected date columns are missing
    if "Start Date" not in df.columns and "End Date" not in df.columns:
        raise HTTPException(status_code=400, detail="This file does not contain Start Date/End Date columns.")

    filtered = df.copy()
    start_series = pd.to_datetime(filtered.get("Start Date"), errors="coerce")
    end_series = pd.to_datetime(filtered.get("End Date"), errors="coerce")

    if start_date:
        try:
            start_cutoff = pd.to_datetime(start_date, errors="raise")
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD.")
        # Include when Start Date exists and is >= cutoff; fallback to End Date when Start Date missing
        start_effective = start_series.fillna(end_series)
        filtered = filtered[start_effective >= start_cutoff]

    if end_date:
        try:
            end_cutoff = pd.to_datetime(end_date, errors="raise")
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD.")
        # Include when End Date exists and is <= cutoff; fallback to Start Date when End Date missing
        end_effective = end_series.fillna(start_series)
        filtered = filtered[end_effective <= end_cutoff]

    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            filtered.to_excel(writer, index=False)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not build filtered Excel: {e}")
    output.seek(0)

    suffix = []
    if start_date:
        suffix.append(f"from_{start_date}")
    if end_date:
        suffix.append(f"to_{end_date}")
    name_suffix = "_".join(suffix) if suffix else "all"
    out_name = f"{Path(file_name).stem}_filtered_{name_suffix}.xlsx"
    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


@app.get("/download")
async def download_list(user: dict = Depends(get_current_user)):
    """List available output Excels with download links."""
    files = []
    user_id = _user_id_from_claims(user)
    ups = user_paths(user_id)
    for p in [ups["excel_sow_path"], ups["excel_invoice_path"]]:
        if p.exists():
            files.append({"file": p.name, "url": f"/download/{p.name}"})
    if not files:
        raise HTTPException(status_code=404, detail="No output files yet; run /extract-sow first.")
    return {"available_files": files}


@app.get("/state")
async def get_state(user: dict = Depends(get_current_user)):
    """Return current extraction state (processed count / path) for debugging."""
    from state_manager import get_state_summary
    user_id = _user_id_from_claims(user)
    spath = user_paths(user_id)["state_path"]
    summary = get_state_summary(state_path=spath)
    summary["state_path"] = str(spath)
    return summary


@app.get("/state/by-type")
async def get_state_by_type(doc_type: str = "sow", user: dict = Depends(get_current_user)):
    """
    Return list of processed files filtered by doc_type.
    Query: ?doc_type=sow or ?doc_type=invoice
    """
    from state_manager import get_processed_items_by_type
    user_id = _user_id_from_claims(user)
    spath = user_paths(user_id)["state_path"]
    items = get_processed_items_by_type(doc_type, state_path=spath)
    return {"doc_type": doc_type, "count": len(items), "items": items}


@app.get("/debug/pdf/download")
async def debug_pdf_download(item_id: str, request: Request, user: dict = Depends(get_current_user)):
    """
    Download a specific PDF from SharePoint by its item ID so you can open it locally.
    Usage: GET /debug/pdf/download?item_id=01JG6ORK...
    Find item IDs from the Excel (or from /debug/pdf/inspect).
    """
    if not is_sharepoint_configured():
        raise HTTPException(status_code=400, detail="SharePoint not configured")
    try:
        user_id = _user_id_from_claims(user)
        ctx = _load_user_context(user_id)
        access_token = get_current_access_token(request)
        loop = asyncio.get_event_loop()
        pdf_bytes = await loop.run_in_executor(
            None,
            lambda: download_file_bytes(
                item_id,
                drive_id=(ctx.get("drive_id") or "").strip() or None,
                access_token=access_token,
                site_url=ctx.get("site_url", ""),
            ),
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {e}")
    from fastapi.responses import Response
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={item_id}.pdf"},
    )


@app.get("/debug/pdf/inspect")
async def debug_pdf_inspect(item_id: str, request: Request, user: dict = Depends(get_current_user)):
    """
    Download a PDF from SharePoint and run full debug: extract text, show preview, run RAG, show per-field results.
    Usage: GET /debug/pdf/inspect?item_id=01JG6ORK...
    Returns: text_length, text_preview (first 2000 chars), text_empty flag, fields dict, and diagnosis.
    """
    if not is_sharepoint_configured():
        raise HTTPException(status_code=400, detail="SharePoint not configured")
    try:
        user_id = _user_id_from_claims(user)
        ctx = _load_user_context(user_id)
        access_token = get_current_access_token(request)
        loop = asyncio.get_event_loop()
        pdf_bytes = await loop.run_in_executor(
            None,
            lambda: download_file_bytes(
                item_id,
                drive_id=(ctx.get("drive_id") or "").strip() or None,
                access_token=access_token,
                site_url=ctx.get("site_url", ""),
            ),
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {e}")
    try:
        result = await loop.run_in_executor(None, lambda: debug_pdf_bytes(pdf_bytes, file_name=item_id))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Debug failed: {e}")
    return result


@app.get("/debug/pdf/list-problems")
async def debug_pdf_list_problems(user: dict = Depends(get_current_user)):
    """
    Read both output Excels and list rows with missing/empty fields or errors.
    Helps identify which PDFs need investigation.
    """
    results = {}
    user_id = _user_id_from_claims(user)
    ups = user_paths(user_id)
    for label, excel_path, check_fields in [
        ("sow", ups["excel_sow_path"], SOW_FIELDS),
        ("invoice", ups["excel_invoice_path"], INVOICE_FIELDS),
    ]:
        if not excel_path.exists():
            results[label] = {"file": excel_path.name, "total_rows": 0, "problem_count": 0, "problems": []}
            continue
        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            results[label] = {"file": excel_path.name, "error": str(e)}
            continue
        problems = []
        for _, row in df.iterrows():
            missing = [f for f in check_fields if not _normalize_field_value(str(row.get(f, "")))]
            has_error = bool(str(row.get("Error", "")).strip())
            if missing or has_error:
                problems.append({
                    "Filename": row.get("Filename", ""),
                    "SharePoint URL": row.get(_URL_KEY, ""),
                    "missing_fields": missing,
                    "error": str(row.get("Error", "")).strip() or None,
                })
        results[label] = {
            "file": excel_path.name,
            "total_rows": len(df),
            "problem_count": len(problems),
            "problems": problems,
        }
    return results


@app.get("/api/ui-version")
async def ui_version():
    """Debug: confirm this process serves the v2 dashboard and root ai_processor extraction."""
    return {
        "dashboard": "v2",
        "extraction_module": "ai_processor (single-shot)",
        "hint": "Single root ai_processor.py. Open GET / and look for 'UI v2'.",
    }


@app.get("/", response_class=HTMLResponse)
async def read_index(request: Request):
    """Serve production-style dashboard UI."""
    if not request.session.get("user"):
        return RedirectResponse(url="/auth/login")
    return HTMLResponse(
        media_type="text/html; charset=utf-8",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        },
        content="""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>SOW & Invoice Extraction Intelligence</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <style>
    /* Fallback when Tailwind CDN is blocked (e.g. corporate VDI) */
    html, body { margin: 0; font-family: system-ui, -apple-system, Segoe UI, Roboto, sans-serif; background: #f8fafc; color: #0f172a; }
    a { color: #1e40af; }
    .fallback-card { background: #fff; border: 1px solid #e2e8f0; border-radius: 0.75rem; padding: 1rem; box-shadow: 0 1px 2px rgba(0,0,0,.05); }
  </style>
</head>
<body class="bg-slate-50 text-slate-900" style="background:#f8fafc;color:#0f172a;">
  <div class="max-w-7xl mx-auto px-4 py-6">
    <header class="flex flex-col md:flex-row md:items-center md:justify-between gap-3">
      <div>
        <h1 class="text-2xl md:text-3xl font-bold">SOW &amp; Invoice Extraction Intelligence</h1>
        <p class="text-sm text-slate-600 mt-1">Production dashboard for SharePoint document processing.</p>
        <p class="text-xs text-indigo-700 font-semibold mt-1" id="uiBuildMarker">UI v2 · single-shot extraction · auto-refresh enabled</p>
      </div>
      <div id="systemBadge" class="inline-flex items-center gap-2 rounded-full px-3 py-1 text-sm font-medium bg-emerald-100 text-emerald-700">
        <span class="w-2 h-2 rounded-full bg-emerald-500"></span>
        System Online
      </div>
    </header>

    <section class="mt-6 grid grid-cols-1 lg:grid-cols-3 gap-4">
      <div class="lg:col-span-2 rounded-xl bg-white border border-slate-200 p-4 shadow-sm">
        <h2 class="text-lg font-semibold mb-3">Control Center</h2>
        <div class="flex flex-wrap items-center gap-2">
          <button id="toggleBtn" onclick="toggleStartStop()" class="px-4 py-2 rounded-md text-white bg-emerald-600 hover:bg-emerald-700">
            Start Extraction
          </button>
          <button onclick="processNow()" class="px-4 py-2 rounded-md text-white bg-indigo-600 hover:bg-indigo-700">
            Process New Files Now
          </button>
          <label class="inline-flex items-center gap-2 text-sm text-slate-700">
            <input id="forceReprocess" type="checkbox" class="rounded border-slate-300" />
            Force reprocess
          </label>
          <button onclick="runScan()" class="px-3 py-2 rounded-md border border-slate-300 hover:bg-slate-100 text-sm">
            Scan
          </button>
          <button onclick="refreshAll()" class="px-3 py-2 rounded-md border border-slate-300 hover:bg-slate-100 text-sm">
            Refresh
          </button>
          <button onclick="logoutUser()" class="px-3 py-2 rounded-md border border-slate-300 hover:bg-slate-100 text-sm">
            Logout
          </button>
        </div>
        <div class="mt-3 grid grid-cols-1 md:grid-cols-3 gap-2">
          <input id="ctxSiteUrl" type="text" placeholder="SharePoint Site URL (https://.../sites/YourSite)" class="border rounded-md px-2 py-2 text-sm" />
          <input id="ctxDrivePath" type="text" placeholder="Drive folder path (e.g. Shared Documents/Contracts)" class="border rounded-md px-2 py-2 text-sm" />
          <input id="ctxDriveId" type="text" placeholder="Optional Drive ID" class="border rounded-md px-2 py-2 text-sm" />
        </div>
        <div class="mt-2">
          <button onclick="saveSharePointContext()" class="px-3 py-2 rounded-md bg-slate-900 text-white hover:bg-slate-700 text-sm">
            Save SharePoint Context
          </button>
        </div>
        <div id="scanResult" class="text-sm text-slate-600 mt-3"></div>
      </div>

      <div class="rounded-xl bg-white border border-slate-200 p-4 shadow-sm">
        <h2 class="text-lg font-semibold mb-3">Download Center</h2>
        <div class="space-y-2">
          <a class="block w-full text-center px-3 py-2 rounded-md bg-slate-900 text-white hover:bg-slate-700" href="/download/license_metrics.xlsx">Download license_metrics.xlsx</a>
          <a class="block w-full text-center px-3 py-2 rounded-md bg-slate-900 text-white hover:bg-slate-700" href="/download/contract_metrics.xlsx">Download contract_metrics.xlsx</a>
        </div>
        <div class="mt-4 border-t pt-3">
          <div class="text-sm font-medium mb-2">Filtered Download (Date Range)</div>
          <label class="block text-xs text-slate-600 mb-1">File</label>
          <select id="filterFile" class="w-full border rounded-md px-2 py-2 text-sm">
            <option value="license_metrics.xlsx">license_metrics.xlsx</option>
            <option value="contract_metrics.xlsx">contract_metrics.xlsx</option>
          </select>
          <div class="grid grid-cols-2 gap-2 mt-2">
            <div>
              <label class="block text-xs text-slate-600 mb-1">Start Date (from)</label>
              <input id="filterStartDate" type="date" class="w-full border rounded-md px-2 py-2 text-sm" />
            </div>
            <div>
              <label class="block text-xs text-slate-600 mb-1">End Date (to)</label>
              <input id="filterEndDate" type="date" class="w-full border rounded-md px-2 py-2 text-sm" />
            </div>
          </div>
          <button onclick="downloadFiltered()" class="w-full mt-3 px-3 py-2 rounded-md bg-indigo-600 text-white hover:bg-indigo-700 text-sm">
            Download Filtered Excel
          </button>
        </div>
        <div id="downloadStatus" class="text-xs text-slate-500 mt-3"></div>
      </div>
    </section>

    <section class="mt-4 grid grid-cols-1 md:grid-cols-2 xl:grid-cols-4 gap-4">
      <div class="rounded-xl bg-white border border-slate-200 p-4 shadow-sm">
        <div class="text-sm text-slate-500">Total Files Processed</div>
        <div id="statTotal" class="text-2xl font-bold mt-1">0</div>
      </div>
      <div class="rounded-xl bg-white border border-slate-200 p-4 shadow-sm">
        <div class="text-sm text-slate-500">SOWs Identified</div>
        <div id="statSow" class="text-2xl font-bold mt-1">0</div>
      </div>
      <div class="rounded-xl bg-white border border-slate-200 p-4 shadow-sm">
        <div class="text-sm text-slate-500">Invoices Identified</div>
        <div id="statInv" class="text-2xl font-bold mt-1">0</div>
      </div>
      <div class="rounded-xl bg-white border border-slate-200 p-4 shadow-sm">
        <div class="text-sm text-slate-500">Errors / Problems</div>
        <div id="statErr" class="text-2xl font-bold mt-1">0</div>
      </div>
    </section>

    <section class="mt-4 rounded-xl bg-white border border-slate-200 p-4 shadow-sm">
      <div class="flex flex-col md:flex-row md:items-center md:justify-between gap-2">
        <h2 class="text-lg font-semibold">Processing Overview</h2>
        <div id="statusHint" class="text-xs text-slate-500">Waiting for status...</div>
      </div>
      <div class="mt-3 grid grid-cols-1 md:grid-cols-2 xl:grid-cols-4 gap-3">
        <div class="rounded-lg border border-slate-200 bg-slate-50 p-3">
          <div class="text-xs text-slate-500">Current Phase</div>
          <div id="overviewPhase" class="text-sm font-semibold text-slate-900 mt-1">-</div>
        </div>
        <div class="rounded-lg border border-slate-200 bg-slate-50 p-3">
          <div class="text-xs text-slate-500">Current File</div>
          <div id="overviewFile" class="text-sm font-semibold text-slate-900 mt-1 break-all">-</div>
        </div>
        <div class="rounded-lg border border-slate-200 bg-slate-50 p-3">
          <div class="text-xs text-slate-500">Processed This Run</div>
          <div id="overviewProcessed" class="text-sm font-semibold text-slate-900 mt-1">0</div>
        </div>
        <div class="rounded-lg border border-slate-200 bg-slate-50 p-3">
          <div class="text-xs text-slate-500">Total To Process</div>
          <div id="overviewTotalToProcess" class="text-sm font-semibold text-slate-900 mt-1">discovering...</div>
        </div>
      </div>
      <div class="mt-3 grid grid-cols-1 md:grid-cols-2 gap-3">
        <div class="rounded-lg border border-slate-200 bg-slate-50 p-3">
          <div class="text-xs text-slate-500">Stop Requested</div>
          <div id="overviewStopRequested" class="text-sm font-semibold text-slate-900 mt-1">No</div>
        </div>
        <div class="rounded-lg border border-slate-200 bg-slate-50 p-3">
          <div class="text-xs text-slate-500">Last Error</div>
          <div id="overviewLastError" class="text-sm font-semibold text-slate-900 mt-1 break-all">None</div>
        </div>
      </div>
    </section>

  </div>

  <div id="toastHost" class="fixed right-4 top-4 z-50 space-y-2"></div>

  <div class="hidden">
    <input id="apiBase" value="" />
  </div>

  <script>
    const uiState = {
      pollTimer: null,
      wasActive: false,
    };

    function buildUrl(path) {
      const base = (document.getElementById("apiBase").value || "").trim();
      if (!base) return path;
      return base.replace(/\\/+$/, "") + path;
    }

    function showToast(msg, type) {
      const host = document.getElementById("toastHost");
      const el = document.createElement("div");
      const color = type === "error" ? "bg-rose-600" : "bg-emerald-600";
      el.className = `${color} text-white px-4 py-2 rounded-md shadow`;
      el.textContent = msg;
      host.appendChild(el);
      setTimeout(() => el.remove(), 4500);
    }

    async function apiCall(path, method, body) {
      let res;
      try {
        res = await fetch(buildUrl(path), {
          method: method || "GET",
          headers: body ? { "Content-Type": "application/json" } : undefined,
          body: body ? JSON.stringify(body) : undefined,
          mode: "cors",
        });
      } catch (e) {
        showToast("Network/CORS error. Check backend URL and CORS settings.", "error");
        throw e;
      }
      const text = await res.text();
      let payload = text;
      try { payload = JSON.parse(text); } catch (e) {}
      if (!res.ok) {
        if (res.status === 401) {
          window.location.href = buildUrl("/auth/login");
          throw new Error("Unauthorized");
        }
        throw new Error(typeof payload === "string" ? payload : JSON.stringify(payload, null, 2));
      }
      return payload;
    }

    async function loadSharePointContext() {
      try {
        const ctx = await apiCall("/sharepoint/context");
        document.getElementById("ctxSiteUrl").value = ctx.site_url || "";
        document.getElementById("ctxDrivePath").value = ctx.drive_path || "";
        document.getElementById("ctxDriveId").value = ctx.drive_id || "";
      } catch (err) {
        // context may not exist yet; ignore
      }
    }

    async function saveSharePointContext() {
      const body = {
        site_url: document.getElementById("ctxSiteUrl").value.trim(),
        drive_path: document.getElementById("ctxDrivePath").value.trim(),
        drive_id: document.getElementById("ctxDriveId").value.trim(),
      };
      try {
        const res = await apiCall("/sharepoint/context", "POST", body);
        showToast(res.validation || "SharePoint context saved", "ok");
      } catch (err) {
        showToast("Save context failed: " + err.message, "error");
      }
    }

    async function logoutUser() {
      try {
        await apiCall("/auth/logout", "POST");
      } catch (e) {}
      window.location.href = buildUrl("/auth/login");
    }

    function setSystemBadge(running) {
      const badge = document.getElementById("systemBadge");
      if (running) {
        badge.className = "inline-flex items-center gap-2 rounded-full px-3 py-1 text-sm font-medium bg-amber-100 text-amber-700";
        badge.innerHTML = '<span class="w-2 h-2 rounded-full bg-amber-500 animate-pulse"></span>Processing...';
      } else {
        badge.className = "inline-flex items-center gap-2 rounded-full px-3 py-1 text-sm font-medium bg-emerald-100 text-emerald-700";
        badge.innerHTML = '<span class="w-2 h-2 rounded-full bg-emerald-500"></span>System Online';
      }
    }

    function setToggleButton(running) {
      const btn = document.getElementById("toggleBtn");
      if (running) {
        btn.textContent = "Stop Extraction";
        btn.className = "px-4 py-2 rounded-md text-white bg-rose-600 hover:bg-rose-700";
      } else {
        btn.textContent = "Start Extraction";
        btn.className = "px-4 py-2 rounded-md text-white bg-emerald-600 hover:bg-emerald-700";
      }
    }

    function startAutoPolling() {
      if (uiState.pollTimer) return;
      uiState.pollTimer = setInterval(refreshStatus, 3000);
    }

    function stopAutoPolling() {
      if (!uiState.pollTimer) return;
      clearInterval(uiState.pollTimer);
      uiState.pollTimer = null;
    }

    async function refreshStatus() {
      try {
        const data = await apiCall("/extract-sow/status");
        setSystemBadge(!!data.running);
        setToggleButton(!!data.running);
        document.getElementById("statErr").textContent = data.last_error ? 1 : 0;

        const total = (data.total_to_process && data.total_to_process > 0) ? data.total_to_process : "discovering...";
        const phase = data.phase || "idle";
        const current = data.current_file || "-";
        const err = data.last_error || "None";
        const stop = data.stop_requested ? "Yes" : "No";

        document.getElementById("overviewPhase").textContent = phase;
        document.getElementById("overviewFile").textContent = current;
        document.getElementById("overviewProcessed").textContent = data.processed_this_run || 0;
        document.getElementById("overviewTotalToProcess").textContent = total;
        document.getElementById("overviewStopRequested").textContent = stop;
        document.getElementById("overviewLastError").textContent = err;
        document.getElementById("statusHint").textContent = `Last updated: ${new Date().toLocaleTimeString()}`;

        const isActive = !!data.running || !!data.stop_requested;
        if (isActive) {
          startAutoPolling();
        } else {
          if (uiState.wasActive) {
            await Promise.all([refreshStateCards(), refreshDownloads()]);
          }
          stopAutoPolling();
        }
        uiState.wasActive = isActive;
      } catch (err) {
        document.getElementById("statusHint").textContent = "Status error: " + err.message;
        setSystemBadge(false);
        setToggleButton(false);
        stopAutoPolling();
      }
    }

    async function refreshStateCards() {
      try {
        const st = await apiCall("/state");
        const by = st.by_doc_type || {};
        document.getElementById("statTotal").textContent = st.processed_count || 0;
        document.getElementById("statSow").textContent = by.sow || 0;
        document.getElementById("statInv").textContent = by.invoice || 0;
      } catch (err) {
        showToast("Failed to fetch /state", "error");
      }
    }

    async function refreshDownloads() {
      try {
        const data = await apiCall("/download");
        const files = (data && data.available_files) ? data.available_files : [];
        const names = files.map(f => f.file).join(", ");
        document.getElementById("downloadStatus").textContent = files.length ? `Available: ${names}` : "No outputs yet.";
      } catch (err) {
        document.getElementById("downloadStatus").textContent = "No outputs yet.";
      }
    }

    function downloadFiltered() {
      const file = document.getElementById("filterFile").value;
      const start = document.getElementById("filterStartDate").value;
      const end = document.getElementById("filterEndDate").value;
      const params = new URLSearchParams();
      if (start) params.set("start_date", start);
      if (end) params.set("end_date", end);
      const url = buildUrl(`/download/filtered/${encodeURIComponent(file)}${params.toString() ? "?" + params.toString() : ""}`);
      window.location.href = url;
    }

    async function startJob() {
      try {
        startAutoPolling();
        const data = await apiCall("/extract-sow/start", "POST");
        showToast(data.message || "Started extraction", "ok");
        await refreshStatus();
      } catch (err) {
        stopAutoPolling();
        showToast("Start failed: " + err.message, "error");
      }
    }

    async function stopJob() {
      try {
        startAutoPolling();
        const data = await apiCall("/extract-sow/stop", "POST");
        showToast(data.message || "Stop requested", "ok");
        await refreshStatus();
        setTimeout(refreshDownloads, 1500);
      } catch (err) {
        showToast("Stop failed: " + err.message, "error");
      }
    }

    async function processNow() {
      const force = document.getElementById("forceReprocess").checked;
      try {
        const path = force ? "/extract-sow/?force_reprocess=true" : "/extract-sow/";
        const data = await apiCall(path, "POST");
        showToast((data && data.message) || "Manual processing complete", "ok");
        await refreshAll();
      } catch (err) {
        showToast("Manual trigger failed: " + err.message, "error");
      }
    }

    async function toggleStartStop() {
      const data = await apiCall("/extract-sow/status");
      if (data.running) return stopJob();
      return startJob();
    }

    async function runScan() {
      try {
        const data = await apiCall("/extract-sow/scan");
        document.getElementById("scanResult").textContent =
          `Scan: total=${data.total}, new=${data.new_count}, changed=${data.changed_count}, up_to_date=${data.up_to_date_count}, to_process=${data.to_process}`;
        showToast("Scan completed", "ok");
      } catch (err) {
        showToast("Scan failed: " + err.message, "error");
      }
    }

    async function refreshAll() {
      try {
        await Promise.all([
          refreshStatus(),
          refreshStateCards(),
          refreshDownloads(),
        ]);
      } catch (e) {
        console.error(e);
        showToast("Refresh failed: " + (e && e.message ? e.message : String(e)), "error");
      }
    }

    loadSharePointContext();
    refreshAll();
  </script>
</body>
</html>
""")


_start_monitor_if_enabled()


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "test":
        import pytest
        sys.exit(pytest.main(["-v", "tests/test_main.py"]))
    import os
    import uvicorn
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
